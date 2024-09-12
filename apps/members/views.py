import datetime
import logging

from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator, PageNotAnInteger, EmptyPage
from django.db.models import Q
from django.http import HttpResponse
from django.shortcuts import render, get_object_or_404
from django.template.loader import get_template
from django.views.decorators.http import require_GET
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Border, Side
from openpyxl.utils import get_column_letter
from xhtml2pdf import pisa

from .models import Membros
from ..pastorais.models import Pastorais


@login_required
@require_GET
def members(request):
    logging.info('Listando membros...')
    
    members_list = Membros.objects.all().prefetch_related('pastorais').select_related('profession').order_by("nome_completo")
    
    filter_params = request.GET.get("search", "")
    if filter_params:
        members_list = members_list.filter(
            Q(nome_completo__icontains=filter_params)
        )
    else:
        members_list = members_list.order_by("nome_completo")

    paginator = Paginator(members_list, 10)
    page = request.GET.get("page", 1)
    try:
        members_list = paginator.page(page)
    except PageNotAnInteger as e:
        logging.error('Numero de pagina não e int: %s', str(e))
        members_list = paginator.page(1)
    except EmptyPage as e:
        logging.error('Pagina Vazia: %s', str(e))
        members_list = paginator.page(paginator.num_pages)

    context = {
        "members_list": members_list,
        "paginator": paginator,
        "page": page,
    }
    logging.info("Lista de membros carregada com sucesso!")
    return render(request, "members/list_members.html", context)


@login_required
@require_GET
def member_details(request, pk):
    membro = get_object_or_404(Membros, pk=pk)

    context = {
        'membro': membro,
    }

    return render(request, "members/member_details.html", context)


@login_required
@require_GET
def export_pdf(request):
    members_list = Membros.objects.all().order_by("nome_completo")
    date = datetime.date.today()
    template_path = 'members/exports/export_pdf.html'
    context = {
        'members_list': members_list,
        'date': date
    }

    colunas_selecionadas = request.GET.getlist('colunas')
    if colunas_selecionadas:
        context['colunas_selecionadas'] = colunas_selecionadas
    else:
        context['colunas_selecionadas'] = ['Selecionar Tudo']

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="lista_de_membros_cadastrados.pdf"'

    template = get_template(template_path)
    table_data = create_table_data_pdf(context['members_list'], context.get('colunas_selecionadas'))
    context['table_data'] = table_data
    html = template.render(context)

    pisa_status = pisa.CreatePDF(html, dest=response)

    if pisa_status.err:
        return HttpResponse('Tivemos alguns erros <pre>' + html + '</pre>')

    return response


def create_table_data_pdf(members_list, colunas_selecionadas):
    opcoes_pastorais = Pastorais.pastoral_choices
    opcoes_elos = Membros.elo_choices
    headers = ['Selecionar Tudo', 'Nome', 'Data de Nascimento', 'Telefone', 'Celular', 'Orientador', 'Pastoral',
               'Elo']

    if 'Selecionar Tudo' in colunas_selecionadas:
        colunas_selecionadas.remove('Selecionar Tudo')

    if colunas_selecionadas:
        headers = [header for header in headers if header in colunas_selecionadas]

    data = []
    for member in members_list:
        member_data = []
        for header in headers:
            if header == 'Nome':
                member_data.append(member.nome_completo if member.nome_completo is not None else 'Não Informado')
            elif header == 'Data de Nascimento':
                member_data.append(
                    member.dt_nascimento.strftime('%d/%m/%Y') if member.dt_nascimento else 'Não Informado')
            elif header == 'Telefone':
                member_data.append(member.telefone if member.telefone is not None else 'Não Informado')
            elif header == 'Celular':
                member_data.append(member.celular if member.celular is not None else 'Não Informado')
            elif header == 'Orientador':
                member_data.append(member.mentor if member.mentor is not None else 'Não Informado')
            elif header == 'Pastoral':
                pastoral_nomes = [pastoral.nome for pastoral in member.pastorais.all() if pastoral.nome[0] is not None]
                pastoral_descriptions = [choice[1] for choice in opcoes_pastorais if choice[0] in pastoral_nomes]
                member_data.append(', '.join(pastoral_descriptions) if pastoral_descriptions else 'Não Informado')
            elif header == 'Elo':
                elo_description = next((choice[1] for choice in opcoes_elos if choice[0] == member.elo), 'Não Informado')
                member_data.append(elo_description)

        data.append(member_data)

    return data


@login_required
@require_GET
def export_excel(request):
    opcoes_pastorais = Pastorais.pastoral_choices
    opcoes_elos = Membros.elo_choices
    members_list = Membros.objects.all()
    colunas_selecionadas = request.GET.getlist('colunas')
    if not colunas_selecionadas or 'Selecionar Tudo' in colunas_selecionadas:
        colunas_selecionadas = ['Nome', 'Data de Nascimento', 'Telefone', 'Celular', 'Orientador', 'Pastoral',
                                'Elo']

    if 'Selecionar Tudo' in colunas_selecionadas:
        colunas_selecionadas.remove('Selecionar Tudo')

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=lista_de_membros_cadastrados.xlsx'

    workbook = Workbook()
    sheet = workbook.active

    headers = colunas_selecionadas
    sheet.append(headers)

    for member in members_list:
        member_data = []
        for header in colunas_selecionadas:
            if header == 'Nome':
                member_data.append(member.nome_completo if member.nome_completo is not None else 'Não Informado')
            elif header == 'Data de Nascimento':
                member_data.append(
                    member.dt_nascimento.strftime('%d/%m/%Y') if member.dt_nascimento else 'Não Informado')
            elif header == 'Telefone':
                member_data.append(member.telefone if member.telefone is not None else 'Não Informado')
            elif header == 'Celular':
                member_data.append(member.celular if member.celular is not None else 'Não Informado')
            elif header == 'Orientador':
                member_data.append(member.mentor if member.mentor is not None else 'Não Informado')
            elif header == 'Pastoral':
                pastoral_nomes = [pastoral.nome for pastoral in member.pastorais.all() if pastoral.nome[0] is not None]
                pastoral_descriptions = [choice[1] for choice in opcoes_pastorais if choice[0] in pastoral_nomes]
                member_data.append(', '.join(pastoral_descriptions) if pastoral_descriptions else 'Não Informado')
            elif header == 'Elo':
                elo_description = next((choice[1] for choice in opcoes_elos if choice[0] == member.elo), 'Não Informado')
                member_data.append(elo_description)
            else:
                member_data.append('')

        sheet.append(member_data)

        for col_num, col_header in enumerate(headers, 1):
            col_letter = get_column_letter(col_num)
            col_width = len(str(col_header)) + 10  
            for row in sheet.iter_rows(min_row=1, max_row=1):
                cell = row[col_num - 1]
                cell.value = col_header
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.font = Font(bold=True)
                cell.border = Border(left=Side(border_style='thin'), right=Side(border_style='thin'))
                sheet.column_dimensions[col_letter].width = col_width

        for row in sheet.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.border = Border(left=Side(border_style='thin'), right=Side(border_style='thin'),
                                     top=Side(border_style='thin'), bottom=Side(border_style='thin'))

    workbook.save(response)
    return response
