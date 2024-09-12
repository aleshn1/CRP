import datetime

from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator, PageNotAnInteger, EmptyPage
from django.db.models import Q
from django.http import HttpResponse
from django.shortcuts import render
from django.template.loader import get_template
from django.views.decorators.http import require_GET
from openpyxl.styles import Alignment, Border, Side, Font
from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook
from xhtml2pdf import pisa

from apps.voluntarios.models import Voluntarios


@login_required
@require_GET
def voluntarios(request):
    context = {}
    voluntarios_list = Voluntarios.objects.all().order_by("nome_completo")

    filter_params = request.GET.get("search", "")
    if filter_params:
        voluntarios_list = voluntarios_list.filter(
            Q(nome_completo__icontains=filter_params) |
            Q(pastorais__icontains=filter_params)
        )
    else:
        voluntarios_list = voluntarios_list.order_by("nome_completo")
    paginator = Paginator(voluntarios_list, 10)
    page = request.GET.get("page", 1)
    try:
        voluntarios_list = paginator.page(page)
    except PageNotAnInteger:
        voluntarios_list = paginator.page(1)
    except EmptyPage:
        voluntarios_list = paginator.page(paginator.num_pages)

    context = {
        "voluntarios_list": voluntarios_list,
        "paginator": paginator,
        "page": page,
    }
    return render(request, 'voluntarios/voluntarios_list.html', context)


@login_required
@require_GET
def export_pdf(request):
    voluntarios_list = Voluntarios.objects.all().order_by("nome_completo")
    date = datetime.date.today()
    template_path = 'voluntarios/exports/export_pdf.html'
    context = {
        'voluntarios_list': voluntarios_list,
        'date': date
    }

    colunas_selecionadas = request.GET.getlist('colunas')
    if colunas_selecionadas:
        context['colunas_selecionadas'] = colunas_selecionadas
    else:
        context['colunas_selecionadas'] = ['Selecionar Tudo']

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="lista_de_voluntarios_cadastrados.pdf"'

    template = get_template(template_path)
    table_data = create_table_data_pdf(context['voluntarios_list'], context.get('colunas_selecionadas'))
    context['table_data'] = table_data
    html = template.render(context)

    pisa_status = pisa.CreatePDF(html, dest=response)

    if pisa_status.err:
        return HttpResponse('Tivemos alguns erros <pre>' + html + '</pre>')

    return response


def create_table_data_pdf(voluntarios_list, colunas_selecionadas):
    opcoes_pastorais = Voluntarios.pastoral_choices
    headers = ['Selecionar Tudo', 'Nome', 'Data de Nascimento', 'Telefone', 'Celular', 'Pastoral']

    if 'Selecionar Tudo' in colunas_selecionadas:
        colunas_selecionadas.remove('Selecionar Tudo')

    if colunas_selecionadas:
        headers = [header for header in headers if header in colunas_selecionadas]

    data = []
    for voluntario in voluntarios_list:
        member_data = []
        for header in headers:
            if header == 'Nome':
                member_data.append(voluntario.nome_completo if voluntario.nome_completo is not None else 'Não Informado')
            elif header == 'Data de Nascimento':
                member_data.append(
                    voluntario.data_nascimento.strftime('%d/%m/%Y') if voluntario.data_nascimento else 'Não Informado')
            elif header == 'Telefone':
                member_data.append(voluntario.telefone if voluntario.telefone is not None else 'Não Informado')
            elif header == 'Celular':
                member_data.append(voluntario.celular if voluntario.celular is not None else 'Não Informado')
            elif header == 'Pastoral':
                pastoral_nomes = [pastoral.nome for pastoral in voluntario.pastorais.all() if pastoral.nome[0] is not None]
                pastoral_descriptions = [choice[1] for choice in opcoes_pastorais if choice[0] in pastoral_nomes]
                member_data.append(', '.join(pastoral_descriptions) if pastoral_descriptions else 'Não Informado')

        data.append(member_data)

    return data


@login_required
@require_GET
def export_excel(request):
    opcoes_pastorais = Voluntarios.pastoral_choices
    voluntarios_list = Voluntarios.objects.all().order_by("nome_completo")
    colunas_selecionadas = request.GET.getlist('colunas')
    if not colunas_selecionadas or 'Selecionar Tudo' in colunas_selecionadas:
        colunas_selecionadas = ['Nome', 'Data de Nascimento', 'Telefone', 'Celular', 'Pastoral',
                                ]

    if 'Selecionar Tudo' in colunas_selecionadas:
        colunas_selecionadas.remove('Selecionar Tudo')

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=lista_de_voluntarios_cadastrados.xlsx'

    workbook = Workbook()
    sheet = workbook.active

    headers = colunas_selecionadas
    sheet.append(headers)

    for voluntario in voluntarios_list:
        member_data = []
        for header in colunas_selecionadas:
            if header == 'Nome':
                member_data.append(voluntario.nome_completo if voluntario.nome_completo is not None else 'Não Informado')
            elif header == 'Data de Nascimento':
                member_data.append(
                    voluntario.data_nascimento.strftime('%d/%m/%Y') if voluntario.data_nascimento else 'Não Informado')
            elif header == 'Telefone':
                member_data.append(voluntario.telefone if voluntario.telefone is not None else 'Não Informado')
            elif header == 'Celular':
                member_data.append(voluntario.celular if voluntario.celular is not None else 'Não Informado')
            elif header == 'Pastoral':
                pastoral_nomes = [pastoral.nome for pastoral in voluntario.pastorais.all() if pastoral.nome[0] is not None]
                pastoral_descriptions = [choice[1] for choice in opcoes_pastorais if choice[0] in pastoral_nomes]
                member_data.append(', '.join(pastoral_descriptions) if pastoral_descriptions else 'Não Informado')
            else:
                member_data.append('')

        sheet.append(member_data)

        # Configurando a largura das colunas
        for col_num, col_header in enumerate(headers, 1):
            col_letter = get_column_letter(col_num)
            col_width = len(str(col_header)) + 10  # Adiciona 2 caracteres para garantir espaço extra
            for row in sheet.iter_rows(min_row=1, max_row=1):
                cell = row[col_num - 1]
                cell.value = col_header
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.font = Font(bold=True)
                cell.border = Border(left=Side(border_style='thin'), right=Side(border_style='thin'))
                sheet.column_dimensions[col_letter].width = col_width

        for row in sheet.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.border = Border(left=Side(border_style='thin'), right=Side(border_style='thin'),
                                     top=Side(border_style='thin'), bottom=Side(border_style='thin'))

    workbook.save(response)
    return response
